# -*- coding: UTF-8 -*-


from xlutils import copy
import xlwt
import xlrd
import openpyxl
import os, sys
import shutil

#
import sys
reload(sys)
sys.setdefaultencoding('utf8')


# test
class Test():

    # excel line config
    _ExportConfigXlsx = os.getcwd() + '/ExportConfig.xlsx'
    _ExportConfig = {}

    # orginal data
    excelFilePath = os.getcwd() + '/../base_data/'
    _col_offset = 6

    # translate data
    _TranslateData = {}
    _TranslatePath = os.getcwd() + '/../Language_data/'
    _TranslateName = "LanguageXlsx"
    _TranslateTitle = ['original', 'change', 'tchinese', 'english', 'korean']

    # out path
    _TranslateOutPath = os.getcwd() + '/../'

    def init(self):
        print("*********************")
        print("init")
        self.LoadExportConfig()
        self.LoadTranslateData()
        print("init done...")
        print("*********************")


    ####TOOL
    # 
    def PathDataToStr(self, pathData):
        pathStr = []
        for _key in pathData:
            path = _key
            for num in pathData[_key]:
                path += "|" + str(num)
            pathStr.append(path)
        return pathStr
    #
    def PathStrToData(self, pathStr):
        data = {}
        for path in pathStr:
            _l = path.split('|')
            _key = _l[0]+ "|" + _l[1]
            data[_key] = []
            for i in range(len(_l)):
                if i > 1 and _l[i] != "": 
                    data[_key].append(int(_l[i]))
        return data


    # English line to num line
    def ColToNum(self, colStr):
        i = 0
        l = len(colStr)
        num = 0
        while i < l:
            num += ((ord(colStr[l - i - 1]) - 64) * (26 ** i))
            i = i + 1
        return num - 1




    def LoadExportConfig(self):
        print("LoadExportConfig...")
        _workBook = xlrd.open_workbook(self._ExportConfigXlsx, 'rb')
        if _workBook:
            _workSheet = _workBook.sheet_by_index(0)
            for _rIndex in range(_workSheet.nrows):
                if _rIndex != 0:
                    _rData = _workSheet.row_values(_rIndex)
                    self._ExportConfig[_rIndex - 1] = _rData


    def LoadTranslateData(self):
        print("LoadTranslateData...")
        self._TranslateData = {}
        _workBook = xlrd.open_workbook(self._TranslatePath + self._TranslateName + ".xls", 'rb')
        if _workBook:
            _workSheet = _workBook.sheet_by_index(0)
            for _rIndex in range(_workSheet.nrows):
                # except title
                if _rIndex != 0:
                    _rData = _workSheet.row_values(_rIndex)
                    _key = _rData[0]
                    # check repeat 
                    if not self._TranslateData.has_key(_key):
                        self._TranslateData[_key] = self.CreateEmptyTranslateItem()
                        _pathStr = []
                        for i in range(len(_rData)):
                            if i < len(self._TranslateTitle):
                                # translate data
                                self._TranslateData[_key][self._TranslateTitle[i]] = _rData[i]
                            else:
                                # path data
                                if _rData[i].strip() != "": _pathStr.append(_rData[i])
                        self._TranslateData[_key]["pathData"] = self.PathStrToData(_pathStr)       

    def CreateEmptyTranslateItem(self):
        o = {}
        for k in self._TranslateTitle:
            o[k] = ""
        o["pathData"] = {}
        return o


    # clear all path
    def ClearTranslateDataPath(self):
        print("clear path...")
        for ol in self._TranslateData:
            self._TranslateData[ol]["pathData"] = {}

    # Add a new data
    def AddTranslateData(self, content, fileName, colName, rowNum):        
        if not self._TranslateData.has_key(content):
            self._TranslateData[content] = self.CreateEmptyTranslateItem()
            self._TranslateData[content]["original"] = content
            
        _pathData = self._TranslateData[content]["pathData"]

        _key = fileName + "|" + colName
        if not _pathData.has_key(_key):
            _pathData[_key] = []

        if not _pathData[_key].__contains__(rowNum):
            _pathData[_key].append(rowNum)
            self._TranslateData[content]["pathData"] = _pathData

    # save to local 
    def WriteTranslateDataToLocal(self):
        print("WriteTranslateDataToLocal...")
        _savePath = self._TranslatePath + self._TranslateName +'.xls'
        _workBook = xlwt.Workbook(encoding='utf-8')
        _workSheet = _workBook.add_sheet(self._TranslateName)

        # title
        _titleIndex = 0
        for _content in self._TranslateTitle:
            _workSheet.write(0, _titleIndex, _content)
            _titleIndex += 1

        # content 
        _rIndex = 1
        for ol in self._TranslateData:
            # translate
            _cIndex = 0
            for _content in self._TranslateTitle:
                _workSheet.write(_rIndex, _cIndex, str(self._TranslateData[ol][_content]))
                _cIndex += 1

            # path
            for p in self.PathDataToStr(self._TranslateData[ol]["pathData"]):
                _workSheet.write(_rIndex, _cIndex, str(p))
                _cIndex += 1

            _rIndex += 1
        
        _workBook.save(_savePath)
        
        print("Write Done")



    def WriteXlsxDataByCol(self, content, path, fileName, colName, rowNums):
        _savePath = path + fileName +'.xlsx'
        # _rBook = xlrd.open_workbook(_savePath)
        _wBook = openpyxl.load_workbook(_savePath)#copy.copy(_rBook)
        _workSheet = _wBook.worksheets[0]

        _colNum = self.ColToNum(colName)
        for row in rowNums:
            _workSheet.cell(row + 1, _colNum + 1, content)
        _wBook.save(_savePath)






    def ExportRead(self, path, fileName, colName):
        _colNum = self.ColToNum(colName)
        _filePath = path + fileName + ".xlsx"
        _workBook = xlrd.open_workbook(_filePath, 'rb')
        _workSheet = _workBook.sheet_by_index(0)
        _cData = _workSheet.col_values(_colNum)
        for _index in range(len(_cData)):
            if _index >= self._col_offset:
                _content = str(_cData[_index]).decode('utf-8')
                if _content and _content != "":
                    self.AddTranslateData(_content, fileName, colName, _index)

    def Export(self):
        print("*********************")
        print("Start Export...")
        self.ClearTranslateDataPath()
        for _index in range(len(self._ExportConfig)):
            _row = self._ExportConfig[_index]
            print("*Read:"+_row[0])
            for i in range(len(_row)):
                if i != 0 and _row[i] != "":
                    self.ExportRead(self.excelFilePath, _row[0], _row[i])
        self.WriteTranslateDataToLocal()  
        print("Export Done!") 
        print("*********************")
                    

    def ReplaceChanges(self):
        # ReExport  
        self.Export()
        # replace
        print("*********************")
        print("replace changes...")
        _replace = 0
        for key in self._TranslateData:
            _tData = self._TranslateData[key]
            _original = _tData['original']
            _change = _tData['change']
            _pathData = _tData['pathData']
            if _change.strip() != "" and _pathData != {} and _original != _change:
                print("*Replace "+ _change +" to "+ _original)
                for fl in _pathData:
                    l = fl.split('|')
                    self.WriteXlsxDataByCol(_change, self.excelFilePath, l[0], l[1], _pathData[fl])
                _replace += 1
                
        # ReExport again
        if _replace > 0:
            self.Export()
        else:
            print("no available changes")
        print("replace done!")
        print("*********************")
       
       


    def DoTranslate(self, language, path, fileName, colName):
        _colNum = self.ColToNum(colName)
        _filePath = path + fileName + ".xlsx"

        _rBook = xlrd.open_workbook(_filePath)
        _rSheet = _rBook.sheet_by_index(0)
        _wBook = openpyxl.load_workbook(_filePath)#copy.copy(_rBook)
        _wSheet = _wBook.worksheets[0]
        print(fileName + " " + str(_colNum))
        _cData = _rSheet.col_values(_colNum)
        for _index in range(len(_cData)):
            if _index >= self._col_offset:
                _content = str(_cData[_index]).decode('utf-8')
                if _content and _content != "":
                    if self._TranslateData.has_key(_content):
                        _translateContent = self._TranslateData[_content][language].decode('utf-8')
                        if _translateContent and _translateContent != "":
                            _wSheet.cell(_index + 1, _colNum + 1, _translateContent)
                        else:
                            print("no ["+ language +"] translate for :" + _content)
        _wBook.save(_filePath)



    def Translate(self, language):
        _absPath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
        _path = _absPath + "/base_data_" + language
        if os.path.exists(_path):
            shutil.rmtree(_path)
        shutil.copytree(_absPath + "/base_data", _path)

        for _index in range(len(self._ExportConfig)):
            _row = self._ExportConfig[_index]
            for i in range(len(_row)):
                if i != 0 and _row[i] != "":
                    self.DoTranslate(language, _path+"/", _row[0], _row[i])



t = Test()
t.init()

if sys.argv[1] == '-e':
    t.Export()
elif sys.argv[1] == '-r':
    t.ReplaceChanges()
elif sys.argv[1] == '-t':
    t.Translate(sys.argv[2])
    pass 
   


# t.Export()
# t.ReplaceChanges()
# t.Translate("english")